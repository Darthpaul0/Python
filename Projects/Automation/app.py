import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    # load the spreadsheet
    wb = xl.load_workbook(filename)

    # select the sheet we'll be working on
    sheet = wb['Hoja1']

    # we can select a cell by using its coordinates
    # cell = sheet.cell(1, 1)

    # we'll iterate through the cells to apply the changes we want
    # we start in row 2 to avoid the column's titles
    # we add one to max_row because range doesn't include the last number
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        # apply discount
        corrected_price = cell.value * 0.9
        # create new column with new prices
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # indicate from where are we getting the data and the range
    values = \
        Reference(sheet,
                  min_row=2, max_row=sheet.max_row,  # range of rows
                  min_col=4, max_col=4)  # range of columns

    # create a chart with the new numbers
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    # save changes
    wb.save(filename)
